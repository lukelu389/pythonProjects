from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request
from datetime import datetime

# Mentee
book = load_workbook("Mentee_Questionnaire_MariHacks.xlsx")
sheet = book["Mentees list"]


class Mentee:
    student_number = 0
    start_date = ""
    end_date = ""
    duration = ""
    program = ""
    gender = ""
    # status indicates if the student is an international student
    status = ""
    # attending highschool
    high_school = ""
    language = ""
    mentor_language = ""
    same_gender_match = False
    extra_activities = []
    goal_extra_activities = []
    other_hobbies = []

    def __init__(self, sn, sd, ed, d, p, g, s, hs, l, ml, sgm, ea, gea, oh):
        self.student_number = sn
        self.start_date = sd
        self.end_date = ed
        self.duration = d
        self.program = p
        self.gender = g
        self.status = s
        self.high_school = hs
        self.language = l
        self.mentor_language = ml
        self.same_gender_match = sgm
        self.extra_activities = ea
        self.goal_extra_activities = gea
        self.other_hobbies = oh

    def rankMentor(self, mentors):
        ranked_mentors = []
        for mentor in mentors:
            rank = 0
            # Rank based on program
            if self.program == mentor.program:
                rank += 5
            # Rank based on gender
            if self.gender == mentor.gender:
                rank += 3
            # Rank based on language
            common_languages = set(self.language).intersection(set(mentor.language))
            rank += len(common_languages)
            # Rank based on hobbies or extracurricular activities
            common_activities = set(self.extra_activities).intersection(set(mentor.extra_activities))
            rank += len(common_activities)
            # Rank based on high school
            if self.high_school == mentor.high_school:
                rank += 2
            ranked_mentors.append((mentor.teacherid, rank))
        # Sort mentors based on rank
        ranked_mentors.sort(key=lambda x: x[1], reverse=True)
        return [teacher_id for teacher_id, _ in ranked_mentors]


# students = []
#
# i = 1
#
# while True:
#     try:
#         t = sheet['A'][i].value
#
#     except IndexError:
#         break
#     try:
#         std_num = sheet["A"][i].value
#         program_s = sheet["B"][i].value
#         program_e = sheet["C"][i].value
#         temp = [sheet["D"][i].value.split(":")]
#         duration = int(temp[0]) * 360 + int(temp[1]) * 60 + int(temp[2])
#         program = sheet["E"][i].value
#         availability = sheet["G"][i].value
#         if availability == 1:
#             availability = True
#         else:
#             availability = False
#
#         gender = sheet["I"][i].value
#         if gender == 0:
#             gender = sheet["H"][i].value
#             if gender == 0:
#                 gender = sheet["J"][i].value
#         status = sheet["K"][i].value
#         if status == 0:
#             status = sheet["L"][i].value
#             if status == 0:
#                 status = ""
#
#         highshool = sheet["N"][i].value
#         languages = []
#         if sheet["O"][i].value == 1: languages.append("English")
#         if sheet["P"][i].value == 1: languages.append("French")
#         if sheet["Q"][i].value == 1:
#             languages.extend(list(sheet["R"][i].value.split()))
#
#         m_languages = []
#         if sheet["S"][i].value == 1: m_languages.append("English")
#         if sheet["T"][i].value == 1: m_languages.append("French")
#         if sheet["U"][i].value == 1:
#             m_languages.extend(list(sheet["V"][i].value.split()))
#
#         same_gender = sheet["W"][i].value
#         if same_gender == 1:
#             same_gender = True
#         else:
#             same_gender = False
#
#         extra_c = []
#         if sheet["X"][i].value == 1:
#             extra_c = list(sheet["Y"][i].value.split())
#
#         f_extra_c = []
#         if sheet["Z"][i].value == 1:
#             f_extra_c = list(sheet["AA"][i].value.split())
#
#         hobbies = []
#         if sheet["AB"][i].value == 1:
#             hobbies = list(sheet["AC"][i].value.split())
#
#         # print(std_num, program_s, program_e, duration, program, gender, status, highshool, languages, m_languages,
#         #       same_gender, extra_c, f_extra_c, hobbies)
#         students.append(
#             Mentee(std_num, program_s, program_e, duration, program, gender, status, highshool, languages, m_languages,
#                    same_gender, extra_c, f_extra_c, hobbies))
#
#     except AttributeError:
#         pass
#     i += 1
#

# Mentor

class Mentor:
    teacherid = ""
    start_date = ""
    end_date = ""
    duration = ""
    program = ""
    gender = ""
    # status indicates if the student is an international student
    status = ""
    # attending highschool
    high_school = ""
    language = []
    student = ""
    same_gender_match = False
    extra_activities = []
    goal_extra_activities = []
    std_amount = 0
    other_hobbies = []
    training_session_time = ""
    students_list = []

    def __init__(self, ti, sd, ed, d, p, g, hs, l, ml, sgm, ea, gea, oh, tst, stda):
        self.teacherid = ti
        self.start_date = sd
        self.end_date = ed
        self.duration = d
        self.program = p
        self.gender = g
        self.high_school = hs
        self.language = l
        self.student = ml
        self.same_gender_match = sgm
        self.extra_activities = ea
        self.goal_extra_activities = gea
        self.other_hobbies = oh
        self.training_session_time = tst
        self.std_amount = stda

    def addStudent(self, std):
        if len(self.students_list) < 2:
            self.students_list.append(std)
        else:
            return False

    def removeStudent(self, std):
        self.students_list.remove(std)


# book = load_workbook("Mentee_Questionnaire_MariHacks.xlsx")
# sheet = book["Mentors list"]
# teachers = []
#
# i = 1
#
# while True:
#     try:
#         t = sheet['A'][i].value
#
#     except IndexError:
#         break
#     try:
#         ti_num = sheet["A"][i].value
#         program_s = sheet["B"][i].value
#         program_e = sheet["C"][i].value
#         temp = [sheet["D"][i].value.split(":")]
#         duration = int(temp[0]) * 360 + int(temp[1]) * 60 + int(temp[2])
#         program = sheet["E"][i].value
#         gender = sheet["G"][i].value
#         if gender == 0:
#             gender = sheet["H"][i].value
#             if gender == 0:
#                 gender = sheet["J"][i].value
#
#         highshool = sheet["K"][i].value
#         languages = []
#         if sheet["L"][i].value == 1: languages.append("English")
#         if sheet["M"][i].value == 1: languages.append("French")
#         if sheet["N"][i].value == 1:
#             languages.extend(list(sheet["O"][i].value.split()))
#
#         m_languages = []
#         if sheet["P"][i].value == 1: m_languages.append("English")
#         if sheet["Q"][i].value == 1: m_languages.append("French")
#         if sheet["R"][i].value == 1:
#             m_languages.extend(list(sheet["S"][i].value.split()))
#
#         std_amount = sheet["T"][i].value
#         same_gender = sheet["U"][i].value
#         if same_gender == 1:
#             same_gender = True
#         else:
#             same_gender = False
#
#         extra_c = []
#         if sheet["V"][i].value == 1:
#             extra_c = list(sheet["W"][i].value.split())
#
#         f_extra_c = []
#         if sheet["X"][i].value == 1:
#             f_extra_c = list(sheet["Y"][i].value.split())
#
#         hobbies = []
#         if sheet["Z"][i].value == 1:
#             hobbies = list(sheet["AA"][i].value.split())
#
#         sessions = sheet["AB"][i].value
#         # print(ti_num, program_s, program_e, duration, program, gender, highshool, languages, m_languages,
#         #            same_gender, extra_c, f_extra_c, hobbies, sessions, std_amount)
#         teachers.append(
#             Mentor(ti_num, program_s, program_e, duration, program, gender, highshool, languages, m_languages,
#                    same_gender, extra_c, f_extra_c, hobbies, sessions, std_amount))
#
#     except AttributeError:
#         pass
#     i += 1
#
# for mentee in students:
#     print(mentee.rankMentor(teachers))

mentees = []
mentors = []

# Load mentees
mentee_book = load_workbook("Mentee_Questionnaire_MariHacks.xlsx")
mentee_sheet = mentee_book["Mentees list"]

for row in mentee_sheet.iter_rows(min_row=2, values_only=True):
    # Convert date strings to datetime objects
    start_date = datetime.strptime(row[1], "%B %dth, %Y %H:%M:%S")
    end_date = datetime.strptime(row[2], "%B %dth, %Y %H:%M:%S")
    # Create Mentee objects and append to mentees list
    mentees.append(
        Mentee(row[0], start_date, end_date, row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11],
               row[12], row[13]))

# Load mentors
mentor_book = load_workbook("Mentee_Questionnaire_MariHacks.xlsx")
mentor_sheet = mentor_book["Mentors list"]

for row in mentor_sheet.iter_rows(min_row=2, values_only=True):
    # Convert date strings to datetime objects
    start_date = datetime.strptime(row[1], "%B %dth, %Y %H:%M:%S")
    end_date = datetime.strptime(row[2],  "%B %dth, %Y %H:%M:%S")
    # Create Mentor objects and append to mentors list
    mentors.append(
        Mentor(row[0], start_date, end_date, row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11],
               row[12], row[13], row[14]))

app = Flask(__name__)


# @app.route("/", methods=["GET", "POST"])
# def index():
#     if request.method == "POST":
#         mentee_id = request.form.get("mentee_id")
#         # Find the mentee with the given ID
#         mentee = next((mentee for mentee in mentees if mentee.student_number == mentee_id), None)
#         if mentee:
#             # Perform ranking logic here
#             ranked_mentors = []  # Replace this with your actual ranking logic
#             return render_template("result.html", mentee=mentee, ranked_mentors=ranked_mentors)
#         else:
#             return render_template("error.html", message="Mentee not found!")
#     return render_template("template.html")
#
#
# if __name__ == "__main__":
#     app.run(debug=True)
