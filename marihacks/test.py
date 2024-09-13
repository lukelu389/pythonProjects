from io import BytesIO
import requests
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request


# Mentee


class Mentee:
    student_number = 0
    start_date = ""
    end_date = ""
    duration = ""
    program = ""
    gender = ""
    # status indicates if the student is an international student
    status = ""
    # attending highschool
    high_school = ""
    language = ""
    mentor_language = ""
    same_gender_match = False
    extra_activities = []
    goal_extra_activities = []
    other_hobbies = []

    def __init__(self, sn, sd, ed, d, p, g, s, hs, l, ml, sgm, ea, gea, oh):
        self.student_number = sn
        self.start_date = sd
        self.end_date = ed
        self.duration = d
        self.program = p
        self.gender = g
        self.status = s
        self.high_school = hs
        self.language = l
        self.mentor_language = ml
        self.same_gender_match = sgm
        self.extra_activities = ea
        self.goal_extra_activities = gea
        self.other_hobbies = oh

    def rankMentor(self, mentors):
        ranked_mentors = []
        for mentor in mentors:
            rank = 0
            # Rank based on program
            if self.program == mentor.program:
                rank += 5
            # Rank based on gender
            if self.gender == mentor.gender:
                rank += 3
            # Rank based on language
            common_languages = set(self.language).intersection(set(mentor.language))
            rank += len(common_languages)
            # Rank based on hobbies or extracurricular activities
            common_activities = set(self.extra_activities).intersection(set(mentor.extra_activities))
            rank += len(common_activities)
            # Rank based on high school
            if self.high_school == mentor.high_school:
                rank += 2
            ranked_mentors.append((mentor.teacherid, rank))
        # Sort mentors based on rank
        ranked_mentors.sort(key=lambda x: x[1], reverse=True)
        return [teacher_id for teacher_id, _ in ranked_mentors]


def load_mentees(file_path):
    students = []
    workbook = load_workbook(file_path)
    sheet = workbook["Mentees list"]

    i = 1

    while True:
        try:
            t = sheet['A'][i].value

        except IndexError:
            break
        try:
            std_num = sheet["A"][i].value
            program_s = sheet["B"][i].value
            program_e = sheet["C"][i].value
            duration = sheet["D"][i].value
            program = sheet["E"][i].value
            availability = sheet["G"][i].value
            if availability == 1:
                availability = True
            else:
                availability = False

            gender = sheet["I"][i].value
            if gender == 0:
                gender = sheet["H"][i].value
                if gender == 0:
                    gender = sheet["J"][i].value
            status = sheet["K"][i].value
            if status == 0:
                status = sheet["L"][i].value
                if status == 0:
                    status = ""

            highshool = sheet["N"][i].value
            languages = []
            if sheet["O"][i].value == 1: languages.append("English")
            if sheet["P"][i].value == 1: languages.append("French")
            if sheet["Q"][i].value == 1:
                languages.extend(list(sheet["R"][i].value.split()))

            m_languages = []
            if sheet["S"][i].value == 1: m_languages.append("English")
            if sheet["T"][i].value == 1: m_languages.append("French")
            if sheet["U"][i].value == 1:
                m_languages.extend(list(sheet["V"][i].value.split()))

            same_gender = sheet["W"][i].value
            if same_gender == 1:
                same_gender = True
            else:
                same_gender = False

            extra_c = []
            if sheet["X"][i].value == 1:
                extra_c = list(sheet["Y"][i].value.split())

            f_extra_c = []
            if sheet["Z"][i].value == 1:
                f_extra_c = list(sheet["AA"][i].value.split())

            hobbies = []
            if sheet["AB"][i].value == 1:
                hobbies = list(sheet["AC"][i].value.split())

            # print(std_num, program_s, program_e, duration, program, gender, status, highshool, languages, m_languages,
            #       same_gender, extra_c, f_extra_c, hobbies)
            students.append(
                Mentee(std_num, program_s, program_e, duration, program, gender, status, highshool, languages,
                       m_languages,
                       same_gender, extra_c, f_extra_c, hobbies))

        except AttributeError:
            pass
        i += 1

    return students


# Mentor

class Mentor:
    teacherid = ""
    start_date = ""
    end_date = ""
    duration = ""
    program = ""
    gender = ""
    # status indicates if the student is an international student
    status = ""
    # attending highschool
    high_school = ""
    language = []
    student = ""
    same_gender_match = False
    extra_activities = []
    goal_extra_activities = []
    std_amount = 0
    other_hobbies = []
    training_session_time = ""

    def __init__(self, ti, sd, ed, d, p, g, hs, l, ml, sgm, ea, gea, oh, tst, stda):
        self.teacherid = ti
        self.start_date = sd
        self.end_date = ed
        self.duration = d
        self.program = p
        self.gender = g
        self.high_school = hs
        self.language = l
        self.student = ml
        self.same_gender_match = sgm
        self.extra_activities = ea
        self.goal_extra_activities = gea
        self.other_hobbies = oh
        self.training_session_time = tst
        self.std_amount = stda


def load_mentors(file_path):
    teachers = []
    workbook = load_workbook(file_path)
    sheet = workbook["Mentors list"]
    i = 1

    while True:
        try:
            t = sheet['A'][i].value

        except IndexError:
            break
        try:
            ti_num = sheet["A"][i].value
            program_s = sheet["B"][i].value
            program_e = sheet["C"][i].value
            duration = sheet["D"][i].value
            program = sheet["E"][i].value

            gender = sheet["G"][i].value
            if gender == 0:
                gender = sheet["H"][i].value
                if gender == 0:
                    gender = sheet["J"][i].value

            highshool = sheet["K"][i].value
            languages = []
            if sheet["L"][i].value == 1: languages.append("English")
            if sheet["M"][i].value == 1: languages.append("French")
            if sheet["N"][i].value == 1:
                languages.extend(list(sheet["O"][i].value.split()))

            m_languages = []
            if sheet["P"][i].value == 1: m_languages.append("English")
            if sheet["Q"][i].value == 1: m_languages.append("French")
            if sheet["R"][i].value == 1:
                m_languages.extend(list(sheet["S"][i].value.split()))

            std_amount = sheet["T"][i].value
            same_gender = sheet["U"][i].value
            if same_gender == 1:
                same_gender = True
            else:
                same_gender = False

            extra_c = []
            if sheet["V"][i].value == 1:
                extra_c = list(sheet["W"][i].value.split())

            f_extra_c = []
            if sheet["X"][i].value == 1:
                f_extra_c = list(sheet["Y"][i].value.split())

            hobbies = []
            if sheet["Z"][i].value == 1:
                hobbies = list(sheet["AA"][i].value.split())

            sessions = sheet["AB"][i].value
            # print(ti_num, program_s, program_e, duration, program, gender, highshool, languages, m_languages,
            #            same_gender, extra_c, f_extra_c, hobbies, sessions, std_amount)
            teachers.append(
                Mentor(ti_num, program_s, program_e, duration, program, gender, highshool, languages, m_languages,
                       same_gender, extra_c, f_extra_c, hobbies, sessions, std_amount))

        except AttributeError:
            pass
        i += 1

    return teachers


app = Flask(__name__)

# Load mentees and mentors

mentees_url = "https://docs.google.com/spreadsheets/d/1IwKx57kCLIWEkWzM6c3v_q8U7uIzOcHKyutBg2l6qT8/export?format=xlsx"
mentors_url = "https://docs.google.com/spreadsheets/d/18gvPwv10XWsWYFGG8h_Nf1FQ4fFp8mchByEBu3Uu2ZM/export?format=xlsx"


def load_google_sheets(url):
    workbook = load_workbook(url)
    return workbook


mentees_workbook = load_google_sheets(mentees_url)
mentors_workbook = load_google_sheets(mentors_url)

mentees = load_mentees(mentees_workbook)
mentors = load_mentors(mentors_workbook)


# Route to handle form submissions
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # Get form data
        form_data = request.form
        # Process form data...
        # For example, you can use the form data to create a new Mentee object
        new_mentee = Mentee(
            form_data["student_number"],
            form_data["start_date"],
            form_data["end_date"],
            form_data["duration"],
            form_data["program"],
            form_data["gender"],
            form_data["status"],
            form_data["high_school"],
            form_data.getlist("language"),
            form_data.getlist("mentor_language"),
            form_data["same_gender_match"],
            form_data.getlist("extra_activities"),
            form_data.getlist("goal_extra_activities"),
            form_data.getlist("other_hobbies")
        )

        ranked_mentors = new_mentee.rankMentor(mentors)
        return render_template("result.html", mentee=new_mentee, ranked_mentors=ranked_mentors)
    return render_template("template.html")


@app.errorhandler(Exception)
def handle_error(error):
    return render_template("error.html", message=str(error))


if __name__ == "__main__":
    app.run(debug=True)
